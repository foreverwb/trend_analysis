"""
Data Import API Routes
Handles Finviz, MarketChameleon, and file imports
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, date
import json
import logging
import io

from ..database import get_db
from ..models import (
    FinvizData, MarketChameleonData, ImportLog, 
    ETFHolding, SectorETF, IndustryETF, SymbolPool, SymbolETFMapping
)
from ..schemas import (
    FinvizImportRequest, FinvizDataItem,
    MarketChameleonImportRequest, MarketChameleonDataItem,
    ImportResponse, ImportLogResponse, HoldingsUpload, HoldingBase
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/import", tags=["Data Import"])


def log_import(db: Session, source: str, etf_symbol: str, count: int, status: str, message: str):
    """Log an import operation"""
    log = ImportLog(
        source=source,
        etf_symbol=etf_symbol,
        record_count=count,
        status=status,
        message=message
    )
    db.add(log)
    db.commit()


# ==================== Finviz Import ====================
def parse_numeric_value(value, default=0):
    """解析数字值，支持字符串格式（如 "1,234,567" 或 "34%"）"""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # 移除逗号、百分号和空格
        cleaned = value.replace(',', '').replace('%', '').replace(' ', '')
        # 处理带 M/B/K 后缀的数字
        if cleaned.endswith('M'):
            return float(cleaned[:-1]) * 1000000
        if cleaned.endswith('B'):
            return float(cleaned[:-1]) * 1000000000
        if cleaned.endswith('K'):
            return float(cleaned[:-1]) * 1000
        try:
            return float(cleaned) if cleaned else default
        except ValueError:
            return default
    return default


@router.post("/finviz", response_model=ImportResponse)
async def import_finviz_data(
    data: FinvizImportRequest,
    db: Session = Depends(get_db)
):
    """Import data from Finviz (JSON format)
    
    支持:
    1. ETF 持仓股票数据 (原有功能)
    2. ETF 自身数据 (如 SPY, XLE 等 ETF 的技术指标)
    """
    try:
        etf_symbol = data.etf_symbol.upper()
        data_date = data.data_date or date.today()
        
        # 检查是否是 ETF 自身的数据（数据中的 Ticker 与 etf_symbol 相同或者是常见的 ETF 代码）
        etf_tickers = {'SPY', 'XLK', 'XLE', 'XLF', 'XLY', 'XLI', 'XLV', 'XLC', 'XLP', 'XLU', 'XLRE', 'XLB',
                       'SOXX', 'SMH', 'IGV', 'XOP', 'XRT', 'KBE', 'IBB', 'XHB', 'XME', 'JETS'}
        
        # 如果数据中包含 ETF 自身，则标记为 ETF 自身数据
        is_etf_self_data = any(
            item.Ticker.upper() in etf_tickers or item.Ticker.upper() == etf_symbol 
            for item in data.data
        )
        
        # Clear existing data for this ETF and date
        db.query(FinvizData).filter(
            FinvizData.etf_symbol == etf_symbol,
            FinvizData.data_date == data_date
        ).delete()
        
        # Insert new data
        count = 0
        for item in data.data:
            ticker = item.Ticker.upper().strip()
            if not ticker:
                continue
            
            # 获取价格（兼容 Price 和 Pirce）
            price = item.Price or item.Pirce or 0
            
            record = FinvizData(
                etf_symbol=etf_symbol,
                ticker=ticker,
                beta=item.Beta,
                atr=item.ATR,
                sma50=item.SMA50,
                sma200=item.SMA200,
                high_52w=item.High_52W,
                rsi=item.RSI,
                price=price,
                volume=item.Volume,
                data_date=data_date
            )
            db.add(record)
            count += 1
        
        db.commit()
        
        # 同步更新标的池状态
        await sync_symbol_pool_after_import(db, etf_symbol, 'finviz')
        
        log_import(db, "finviz", etf_symbol, count, "success", 
                   f"Imported {count} records" + (" (ETF self data)" if is_etf_self_data else ""))
        
        return ImportResponse(
            success=True,
            source="finviz",
            etf_symbol=etf_symbol,
            record_count=count,
            message=f"Successfully imported {count} Finviz records for {etf_symbol}",
            timestamp=datetime.now()
        )
    except Exception as e:
        logger.error(f"Finviz import error: {e}")
        log_import(db, "finviz", data.etf_symbol, 0, "failed", str(e))
        raise HTTPException(status_code=400, detail=str(e))


# ==================== MarketChameleon Import ====================
@router.post("/marketchameleon", response_model=ImportResponse)
async def import_marketchameleon_data(
    data: MarketChameleonImportRequest,
    db: Session = Depends(get_db)
):
    """Import data from MarketChameleon (JSON format)
    
    支持:
    1. ETF 持仓股票数据 (原有功能)
    2. ETF 自身数据 (如 SPY 等 ETF 的期权指标)
    3. 字符串格式的数值 (如 "3,875,171", "55.7%")
    """
    try:
        etf_symbol = data.etf_symbol.upper() if data.etf_symbol else None
        data_date = data.data_date or date.today()
        
        # Clear existing data for this date (and ETF if specified)
        query = db.query(MarketChameleonData).filter(
            MarketChameleonData.data_date == data_date
        )
        if etf_symbol:
            query = query.filter(MarketChameleonData.etf_symbol == etf_symbol)
        query.delete()
        
        count = 0
        for item in data.data:
            symbol = item.symbol.upper().strip()
            if not symbol:
                continue
            
            # 解析各种格式的数值
            record = MarketChameleonData(
                etf_symbol=etf_symbol,
                symbol=symbol,
                rel_notional_to_90d=parse_numeric_value(item.RelNotionalTo90D),
                rel_vol_to_90d=parse_numeric_value(item.RelVolTo90D),
                trade_count=int(parse_numeric_value(item.TradeCount)),
                iv30=parse_numeric_value(item.IV30),
                hv20=parse_numeric_value(item.HV20),
                ivr=parse_numeric_value(item.IVR),
                iv_52w_p=parse_numeric_value(item.IV_52W_P),
                iv30_chg=parse_numeric_value(item.IV30_Chg or item.IV30ChgPct),
                multi_leg_pct=parse_numeric_value(item.MultiLegPct),
                contingent_pct=parse_numeric_value(item.ContingentPct),
                put_pct=parse_numeric_value(item.PutPct),
                call_volume=int(parse_numeric_value(item.CallVolume)),
                put_volume=int(parse_numeric_value(item.PutVolume)),
                data_date=data_date
            )
            db.add(record)
            count += 1
        
        db.commit()
        
        # 同步更新标的池状态
        if etf_symbol:
            await sync_symbol_pool_after_import(db, etf_symbol, 'marketchameleon')
        
        log_import(db, "marketchameleon", etf_symbol, count, "success", f"Imported {count} records")
        
        return ImportResponse(
            success=True,
            source="marketchameleon",
            etf_symbol=etf_symbol,
            record_count=count,
            message=f"Successfully imported {count} MarketChameleon records",
            timestamp=datetime.now()
        )
    except Exception as e:
        logger.error(f"MarketChameleon import error: {e}")
        log_import(db, "marketchameleon", data.etf_symbol, 0, "failed", str(e))
        raise HTTPException(status_code=400, detail=str(e))


async def sync_symbol_pool_after_import(db: Session, etf_symbol: str, source: str):
    """导入数据后同步更新标的池状态
    
    修复 Bug #2.c: 导入数据后"标的池状态明细"Finviz标识有误
    修复: 导入数据后自动更新ETF的广度评分
    """
    from ..models import SymbolPool, FinvizData, MarketChameleonData, SymbolETFMapping, SectorETF, IndustryETF
    from ..services.calculation import CalculationService
    
    calc_service = CalculationService(db)
    
    # 获取该 ETF 下所有标的的最新数据
    if source == 'finviz':
        tickers = db.query(FinvizData.ticker).filter(
            FinvizData.etf_symbol == etf_symbol
        ).distinct().all()
        
        for (ticker,) in tickers:
            symbol = db.query(SymbolPool).filter(SymbolPool.ticker == ticker).first()
            
            # Bug #2.c 修复: 如果标的不存在，创建它
            if not symbol:
                symbol = SymbolPool(
                    ticker=ticker,
                    finviz_status='pending',
                    mc_status='pending',
                    ibkr_status='pending',
                    futu_status='pending'
                )
                db.add(symbol)
                db.flush()
                
                # 创建ETF映射关系
                existing_mapping = db.query(SymbolETFMapping).filter(
                    SymbolETFMapping.ticker == ticker,
                    SymbolETFMapping.etf_symbol == etf_symbol
                ).first()
                
                if not existing_mapping:
                    mapping = SymbolETFMapping(
                        ticker=ticker,
                        etf_symbol=etf_symbol,
                        etf_type='sector' if etf_symbol in ['XLK', 'XLF', 'XLE', 'XLV', 'XLY', 'XLI', 'XLC', 'XLP', 'XLU', 'XLRE', 'XLB'] else 'industry',
                        weight=0,
                        rank=0
                    )
                    db.add(mapping)
            
            # 更新状态
            symbol.finviz_status = 'ready'
            symbol.finviz_last_update = datetime.utcnow()
            # 重新计算完备度
            count = 0
            if symbol.finviz_status == 'ready': count += 1
            if symbol.mc_status == 'ready': count += 1
            if symbol.ibkr_status == 'ready': count += 1
            if symbol.futu_status == 'ready': count += 1
            symbol.completeness = int((count / 4) * 100)
            symbol.updated_at = datetime.utcnow()
        
        # 导入 Finviz 数据后，自动更新 ETF 的广度评分
        finviz_data = db.query(FinvizData).filter(
            FinvizData.etf_symbol == etf_symbol
        ).all()
        
        if finviz_data:
            breadth_score, pct_above_50ma, pct_above_200ma = calc_service.calculate_breadth_score(finviz_data)
            
            # 更新 Sector ETF
            sector_etf = db.query(SectorETF).filter(SectorETF.symbol == etf_symbol).first()
            if sector_etf:
                sector_etf.breadth_score = breadth_score
                sector_etf.pct_above_50ma = pct_above_50ma
                sector_etf.pct_above_200ma = pct_above_200ma
                # 重新计算综合分
                sector_etf.composite_score = calc_service.calculate_etf_composite_score(
                    sector_etf.rel_momentum_score or 0,
                    sector_etf.trend_quality_score or 0,
                    breadth_score,
                    sector_etf.options_score or 0
                )
                sector_etf.updated_at = datetime.utcnow()
            
            # 更新 Industry ETF
            industry_etf = db.query(IndustryETF).filter(IndustryETF.symbol == etf_symbol).first()
            if industry_etf:
                industry_etf.breadth_score = breadth_score
                industry_etf.pct_above_50ma = pct_above_50ma
                industry_etf.pct_above_200ma = pct_above_200ma
                industry_etf.composite_score = calc_service.calculate_etf_composite_score(
                    industry_etf.rel_momentum_score or 0,
                    industry_etf.trend_quality_score or 0,
                    breadth_score,
                    industry_etf.options_score or 0
                )
                industry_etf.updated_at = datetime.utcnow()
    
    elif source == 'marketchameleon':
        symbols = db.query(MarketChameleonData.symbol).filter(
            MarketChameleonData.etf_symbol == etf_symbol
        ).distinct().all()
        
        for (ticker,) in symbols:
            symbol = db.query(SymbolPool).filter(SymbolPool.ticker == ticker).first()
            
            # Bug #2.c 修复: 如果标的不存在，创建它
            if not symbol:
                symbol = SymbolPool(
                    ticker=ticker,
                    finviz_status='pending',
                    mc_status='pending',
                    ibkr_status='pending',
                    futu_status='pending'
                )
                db.add(symbol)
                db.flush()
                
                # 创建ETF映射关系
                existing_mapping = db.query(SymbolETFMapping).filter(
                    SymbolETFMapping.ticker == ticker,
                    SymbolETFMapping.etf_symbol == etf_symbol
                ).first()
                
                if not existing_mapping:
                    mapping = SymbolETFMapping(
                        ticker=ticker,
                        etf_symbol=etf_symbol,
                        etf_type='sector' if etf_symbol in ['XLK', 'XLF', 'XLE', 'XLV', 'XLY', 'XLI', 'XLC', 'XLP', 'XLU', 'XLRE', 'XLB'] else 'industry',
                        weight=0,
                        rank=0
                    )
                    db.add(mapping)
            
            # 更新状态
            symbol.mc_status = 'ready'
            symbol.mc_last_update = datetime.utcnow()
            # 重新计算完备度
            count = 0
            if symbol.finviz_status == 'ready': count += 1
            if symbol.mc_status == 'ready': count += 1
            if symbol.ibkr_status == 'ready': count += 1
            if symbol.futu_status == 'ready': count += 1
            symbol.completeness = int((count / 4) * 100)
            symbol.updated_at = datetime.utcnow()
        
        # 导入 MarketChameleon 数据后，自动更新 ETF 的期权评分
        mc_data = db.query(MarketChameleonData).filter(
            MarketChameleonData.etf_symbol == etf_symbol
        ).all()
        
        if mc_data:
            options_score, options_heat, rel_vol, ivr = calc_service.calculate_options_confirm_score(mc_data)
            
            # 更新 Sector ETF
            sector_etf = db.query(SectorETF).filter(SectorETF.symbol == etf_symbol).first()
            if sector_etf:
                sector_etf.options_score = options_score
                sector_etf.options_heat = options_heat
                sector_etf.rel_vol = rel_vol
                sector_etf.ivr = ivr
                sector_etf.composite_score = calc_service.calculate_etf_composite_score(
                    sector_etf.rel_momentum_score or 0,
                    sector_etf.trend_quality_score or 0,
                    sector_etf.breadth_score or 0,
                    options_score
                )
                sector_etf.updated_at = datetime.utcnow()
            
            # 更新 Industry ETF
            industry_etf = db.query(IndustryETF).filter(IndustryETF.symbol == etf_symbol).first()
            if industry_etf:
                industry_etf.options_score = options_score
                industry_etf.options_heat = options_heat
                industry_etf.rel_vol = rel_vol
                industry_etf.ivr = ivr
                industry_etf.composite_score = calc_service.calculate_etf_composite_score(
                    industry_etf.rel_momentum_score or 0,
                    industry_etf.trend_quality_score or 0,
                    industry_etf.breadth_score or 0,
                    options_score
                )
                industry_etf.updated_at = datetime.utcnow()
    
    db.commit()


# ==================== File Upload ====================
@router.post("/upload/json", response_model=ImportResponse)
async def upload_json_file(
    file: UploadFile = File(...),
    source: str = Form(...),  # 'finviz' or 'marketchameleon'
    etf_symbol: str = Form(...),
    db: Session = Depends(get_db)
):
    """Upload JSON file for import"""
    try:
        content = await file.read()
        data = json.loads(content.decode('utf-8'))
        
        if source == "finviz":
            # Convert to request format
            if isinstance(data, list):
                items = [FinvizDataItem(**item) for item in data]
            elif isinstance(data, dict) and "data" in data:
                items = [FinvizDataItem(**item) for item in data["data"]]
            else:
                raise ValueError("Invalid JSON format")
            
            request = FinvizImportRequest(etf_symbol=etf_symbol, data=items)
            return await import_finviz_data(request, db)
        
        elif source == "marketchameleon":
            if isinstance(data, list):
                items = [MarketChameleonDataItem(**item) for item in data]
            elif isinstance(data, dict) and "data" in data:
                items = [MarketChameleonDataItem(**item) for item in data["data"]]
            else:
                raise ValueError("Invalid JSON format")
            
            request = MarketChameleonImportRequest(etf_symbol=etf_symbol, data=items)
            return await import_marketchameleon_data(request, db)
        
        else:
            raise ValueError(f"Unknown source: {source}")
    
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")
    except Exception as e:
        logger.error(f"File upload error: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/upload/xlsx", response_model=ImportResponse)
async def upload_xlsx_holdings(
    file: UploadFile = File(...),
    etf_type: str = Form(...),  # 'sector' or 'industry'
    etf_symbol: str = Form(...),
    sector_symbol: Optional[str] = Form(None),  # For industry ETF
    data_date: str = Form(...),  # YYYY-MM-DD
    db: Session = Depends(get_db)
):
    """
    Upload XLSX file for ETF holdings
    Extracts 'Ticker' and 'Weight' or 'Weight %' columns
    """
    try:
        import openpyxl
        
        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
        
        # Find column indices
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        ticker_col = None
        weight_col = None
        
        for idx, col_name in enumerate(header_row):
            if col_name and 'ticker' in str(col_name).lower():
                ticker_col = idx
            if col_name and ('weight' in str(col_name).lower()):
                weight_col = idx
        
        if ticker_col is None or weight_col is None:
            raise ValueError("Could not find 'Ticker' and 'Weight' columns in XLSX")
        
        # Parse holdings
        holdings = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ticker = row[ticker_col]
            weight = row[weight_col]
            
            # Validate ticker
            if not ticker or not isinstance(ticker, str):
                continue
            ticker = ticker.strip().upper()
            if not ticker.isalpha():
                continue
            
            # Parse weight
            if weight is None:
                continue
            if isinstance(weight, str):
                weight = float(weight.replace('%', '').strip())
            else:
                weight = float(weight)
            
            holdings.append(HoldingBase(ticker=ticker, weight=weight))
        
        if not holdings:
            raise ValueError("No valid holdings found in XLSX")
        
        # Create upload request
        etf_symbol = etf_symbol.upper()
        parsed_date = datetime.strptime(data_date, "%Y-%m-%d").date()
        
        upload_data = HoldingsUpload(
            etf_symbol=etf_symbol,
            etf_type=etf_type,
            sector_symbol=sector_symbol.upper() if sector_symbol else None,
            data_date=parsed_date,
            holdings=holdings
        )
        
        # Process upload (reuse logic from etf router)
        if etf_type == "sector":
            etf = db.query(SectorETF).filter(SectorETF.symbol == etf_symbol).first()
            if not etf:
                etf = SectorETF(symbol=etf_symbol, name=etf_symbol)
                db.add(etf)
                db.commit()
            
            db.query(ETFHolding).filter(
                ETFHolding.sector_etf_symbol == etf_symbol,
                ETFHolding.data_date == parsed_date
            ).delete()
            
            for holding in holdings:
                h = ETFHolding(
                    etf_type="sector",
                    etf_symbol=etf_symbol,
                    sector_etf_symbol=etf_symbol,
                    ticker=holding.ticker,
                    weight=holding.weight,
                    data_date=parsed_date
                )
                db.add(h)
        else:
            etf = db.query(IndustryETF).filter(IndustryETF.symbol == etf_symbol).first()
            if not etf:
                etf = IndustryETF(
                    symbol=etf_symbol,
                    name=etf_symbol,
                    sector_symbol=sector_symbol.upper() if sector_symbol else None
                )
                db.add(etf)
                db.commit()
            
            db.query(ETFHolding).filter(
                ETFHolding.industry_etf_symbol == etf_symbol,
                ETFHolding.data_date == parsed_date
            ).delete()
            
            for holding in holdings:
                h = ETFHolding(
                    etf_type="industry",
                    etf_symbol=etf_symbol,
                    industry_etf_symbol=etf_symbol,
                    ticker=holding.ticker,
                    weight=holding.weight,
                    data_date=parsed_date
                )
                db.add(h)
        
        db.commit()
        
        log_import(db, "xlsx", etf_symbol, len(holdings), "success", f"Uploaded {len(holdings)} holdings")
        
        return ImportResponse(
            success=True,
            source="xlsx",
            etf_symbol=etf_symbol,
            record_count=len(holdings),
            message=f"Successfully uploaded {len(holdings)} holdings for {etf_symbol}",
            timestamp=datetime.now()
        )
    
    except Exception as e:
        logger.error(f"XLSX upload error: {e}")
        log_import(db, "xlsx", etf_symbol, 0, "failed", str(e))
        raise HTTPException(status_code=400, detail=str(e))


# ==================== Import History ====================
@router.get("/history", response_model=List[ImportLogResponse])
async def get_import_history(
    limit: int = 20,
    source: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get import history"""
    query = db.query(ImportLog)
    if source:
        query = query.filter(ImportLog.source == source)
    
    logs = query.order_by(ImportLog.created_at.desc()).limit(limit).all()
    return logs


@router.delete("/history/{log_id}")
async def delete_import_log(log_id: int, db: Session = Depends(get_db)):
    """Delete an import log entry"""
    log = db.query(ImportLog).filter(ImportLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Import log not found")
    
    db.delete(log)
    db.commit()
    return {"message": "Import log deleted"}


# ==================== Template Download ====================
@router.get("/template/{template_type}")
async def get_import_template(template_type: str):
    """Get JSON template for import"""
    templates = {
        "finviz": {
            "etf_symbol": "XLK",
            "data": [
                {
                    "Ticker": "AAPL",
                    "Beta": 1.2,
                    "ATR": 3.5,
                    "SMA50": 180.0,
                    "SMA200": 175.0,
                    "52W_High": 200.0,
                    "RSI": 55.0,
                    "Price": 185.0,
                    "Volume": 50000000
                }
            ]
        },
        "marketchameleon": {
            "etf_symbol": "XLK",
            "data": [
                {
                    "symbol": "AAPL",
                    "RelNotionalTo90D": 1.5,
                    "RelVolTo90D": 1.3,
                    "TradeCount": 50000,
                    "IV30": 25.5,
                    "HV20": 22.0,
                    "IVR": 65.0,
                    "IV_52W_P": 60.0,
                    "IV30_Chg": 2.5,
                    "MultiLegPct": 35.0,
                    "ContingentPct": 10.0,
                    "PutPct": 40.0,
                    "CallVolume": 100000,
                    "PutVolume": 80000
                }
            ]
        }
    }
    
    if template_type not in templates:
        raise HTTPException(status_code=404, detail=f"Template {template_type} not found")
    
    return templates[template_type]
