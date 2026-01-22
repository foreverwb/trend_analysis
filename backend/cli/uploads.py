#!/usr/bin/env python3
"""
CLI Tool for ETF Holdings Upload

Usage:
    # Sector ETF holdings upload
    python -m backend.cli.uploads -d 2024-01-15 -t sector -a XLK path/to/holdings.xlsx
    python -m backend.cli.uploads -d 2024-01-15 -t sector -a XLK path/to/holdings.csv
    
    # Industry ETF holdings upload
    python -m backend.cli.uploads -d 2024-01-15 -t industry -s XLK -a SOXX path/to/holdings.xlsx
    python -m backend.cli.uploads -d 2024-01-15 -t industry -s XLK -a SOXX path/to/holdings.csv
"""
import argparse
import sys
import os
import csv
from datetime import datetime
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import openpyxl
from sqlalchemy.orm import Session

from backend.database import SessionLocal, init_db
from backend.models import SectorETF, IndustryETF, ETFHolding


VALID_SECTOR_ETFS = ['XLK', 'XLC', 'XLY', 'XLP', 'XLV', 'XLF', 'XLI', 'XLE', 'XLU', 'XLRE', 'XLB']


def parse_xlsx(file_path: str) -> list:
    """
    Parse XLSX file and extract Ticker and Weight columns
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Find column indices
    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    ticker_col = None
    weight_col = None
    
    for idx, col_name in enumerate(header_row):
        if col_name is None:
            continue
        col_lower = str(col_name).lower()
        if 'ticker' in col_lower or 'symbol' in col_lower:
            ticker_col = idx
        if 'weight' in col_lower:
            weight_col = idx
    
    if ticker_col is None:
        raise ValueError("Could not find 'Ticker' or 'Symbol' column in XLSX file")
    if weight_col is None:
        raise ValueError("Could not find 'Weight' or 'Weight %' column in XLSX file")
    
    holdings = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        ticker = row[ticker_col]
        weight = row[weight_col]
        
        # Validate ticker: must be non-empty and contain only English letters
        if not ticker:
            continue
        ticker = str(ticker).strip().upper()
        if not ticker.isalpha():
            print(f"  Skipping invalid ticker: {ticker}")
            continue
        
        # Parse weight
        if weight is None:
            continue
        if isinstance(weight, str):
            weight = float(weight.replace('%', '').strip())
        else:
            weight = float(weight)
        
        holdings.append({'ticker': ticker, 'weight': weight})
    
    return holdings


def parse_csv(file_path: str) -> list:
    """
    Parse CSV file and extract Ticker and Weight columns
    """
    holdings = []
    
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        # Try to detect delimiter
        sample = f.read(4096)
        f.seek(0)
        
        # Check for common delimiters
        if '\t' in sample:
            delimiter = '\t'
        elif ';' in sample:
            delimiter = ';'
        else:
            delimiter = ','
        
        reader = csv.DictReader(f, delimiter=delimiter)
        
        # Find the correct column names (case-insensitive)
        fieldnames_lower = {name.lower(): name for name in reader.fieldnames} if reader.fieldnames else {}
        
        ticker_col = None
        weight_col = None
        
        for lower_name, original_name in fieldnames_lower.items():
            if 'ticker' in lower_name or 'symbol' in lower_name:
                ticker_col = original_name
            if 'weight' in lower_name:
                weight_col = original_name
        
        if ticker_col is None:
            raise ValueError(f"Could not find 'Ticker' or 'Symbol' column in CSV file. Found columns: {list(reader.fieldnames or [])}")
        if weight_col is None:
            raise ValueError(f"Could not find 'Weight' column in CSV file. Found columns: {list(reader.fieldnames or [])}")
        
        print(f"  Using columns: Ticker='{ticker_col}', Weight='{weight_col}'")
        
        for row in reader:
            ticker = row.get(ticker_col, '').strip()
            weight_str = row.get(weight_col, '').strip()
            
            # Validate ticker: must be non-empty and contain only English letters
            if not ticker:
                continue
            ticker = ticker.upper()
            if not ticker.isalpha():
                print(f"  Skipping invalid ticker: {ticker}")
                continue
            
            # Parse weight
            if not weight_str:
                continue
            try:
                weight = float(weight_str.replace('%', '').strip())
            except ValueError:
                print(f"  Skipping invalid weight for {ticker}: {weight_str}")
                continue
            
            holdings.append({'ticker': ticker, 'weight': weight})
    
    return holdings


def parse_file(file_path: str) -> list:
    """
    Parse file based on extension (supports .xlsx and .csv)
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.xlsx':
        return parse_xlsx(file_path)
    elif ext == '.csv':
        return parse_csv(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}. Use .xlsx or .csv")


def upload_holdings(
    db: Session,
    data_date: datetime.date,
    etf_type: str,
    etf_symbol: str,
    sector_symbol: str,
    holdings: list
):
    """Upload holdings to database"""
    etf_symbol = etf_symbol.upper()
    
    if etf_type == 'sector':
        # Ensure sector ETF exists
        etf = db.query(SectorETF).filter(SectorETF.symbol == etf_symbol).first()
        if not etf:
            etf = SectorETF(symbol=etf_symbol, name=etf_symbol)
            db.add(etf)
            db.commit()
        
        # Delete existing holdings for this date
        deleted = db.query(ETFHolding).filter(
            ETFHolding.sector_etf_symbol == etf_symbol,
            ETFHolding.data_date == data_date
        ).delete()
        print(f"  Deleted {deleted} existing holdings")
        
        # Add new holdings
        for h in holdings:
            holding = ETFHolding(
                etf_type='sector',
                etf_symbol=etf_symbol,
                sector_etf_symbol=etf_symbol,
                ticker=h['ticker'],
                weight=h['weight'],
                data_date=data_date
            )
            db.add(holding)
    
    else:  # industry
        sector_symbol = sector_symbol.upper() if sector_symbol else None
        
        # Ensure industry ETF exists
        etf = db.query(IndustryETF).filter(IndustryETF.symbol == etf_symbol).first()
        if not etf:
            etf = IndustryETF(
                symbol=etf_symbol,
                name=etf_symbol,
                sector_symbol=sector_symbol
            )
            db.add(etf)
            db.commit()
        
        # Delete existing holdings for this date
        deleted = db.query(ETFHolding).filter(
            ETFHolding.industry_etf_symbol == etf_symbol,
            ETFHolding.data_date == data_date
        ).delete()
        print(f"  Deleted {deleted} existing holdings")
        
        # Add new holdings
        for h in holdings:
            holding = ETFHolding(
                etf_type='industry',
                etf_symbol=etf_symbol,
                industry_etf_symbol=etf_symbol,
                ticker=h['ticker'],
                weight=h['weight'],
                data_date=data_date
            )
            db.add(holding)
    
    db.commit()


def main():
    parser = argparse.ArgumentParser(
        description='Upload ETF holdings from XLSX or CSV file',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  Sector ETF:
    uploads -d 2024-01-15 -t sector -a XLK holdings.xlsx
    uploads -d 2024-01-15 -t sector -a XLK holdings.csv
    
  Industry ETF:
    uploads -d 2024-01-15 -t industry -s XLK -a SOXX holdings.xlsx
    uploads -d 2024-01-15 -t industry -s XLK -a SOXX holdings.csv
    
Valid Sector ETFs: XLK, XLC, XLY, XLP, XLV, XLF, XLI, XLE, XLU, XLRE, XLB

CSV/XLSX columns required:
  - Ticker (or Symbol): Stock ticker symbol
  - Weight: Holding weight percentage
        """
    )
    
    parser.add_argument(
        '-d', '--date',
        required=True,
        help='Data date in YYYY-MM-DD format'
    )
    parser.add_argument(
        '-t', '--type',
        required=True,
        choices=['sector', 'industry'],
        help='ETF type: sector or industry'
    )
    parser.add_argument(
        '-s', '--sector',
        help='Parent sector symbol (required for industry ETF)'
    )
    parser.add_argument(
        '-a', '--etf',
        required=True,
        help='ETF symbol (e.g., XLK for sector, SOXX for industry)'
    )
    parser.add_argument(
        'file',
        help='Path to XLSX or CSV file containing holdings'
    )
    
    args = parser.parse_args()
    
    # Validate arguments
    try:
        data_date = datetime.strptime(args.date, '%Y-%m-%d').date()
    except ValueError:
        print(f"Error: Invalid date format '{args.date}'. Use YYYY-MM-DD.")
        sys.exit(1)
    
    if args.type == 'sector':
        if args.etf.upper() not in VALID_SECTOR_ETFS:
            print(f"Warning: {args.etf} is not a standard sector ETF")
    
    if args.type == 'industry' and not args.sector:
        print("Error: -s/--sector is required for industry ETF uploads")
        sys.exit(1)
    
    if not os.path.exists(args.file):
        print(f"Error: File not found: {args.file}")
        sys.exit(1)
    
    # Validate file extension
    ext = os.path.splitext(args.file)[1].lower()
    if ext not in ['.xlsx', '.csv']:
        print(f"Error: Unsupported file format '{ext}'. Use .xlsx or .csv")
        sys.exit(1)
    
    # Initialize database
    init_db()
    
    # Parse file
    print(f"\nParsing {args.file}...")
    try:
        holdings = parse_file(args.file)
    except Exception as e:
        print(f"Error parsing file: {e}")
        sys.exit(1)
    
    if not holdings:
        print("Error: No valid holdings found in file")
        sys.exit(1)
    
    print(f"  Found {len(holdings)} valid holdings")
    
    # Upload to database
    print(f"\nUploading to database...")
    print(f"  Date: {data_date}")
    print(f"  Type: {args.type}")
    print(f"  ETF: {args.etf.upper()}")
    if args.sector:
        print(f"  Sector: {args.sector.upper()}")
    
    db = SessionLocal()
    try:
        upload_holdings(
            db=db,
            data_date=data_date,
            etf_type=args.type,
            etf_symbol=args.etf,
            sector_symbol=args.sector,
            holdings=holdings
        )
        print(f"\n✓ Successfully uploaded {len(holdings)} holdings for {args.etf.upper()}")
    except Exception as e:
        print(f"\nError uploading holdings: {e}")
        sys.exit(1)
    finally:
        db.close()


if __name__ == '__main__':
    main()